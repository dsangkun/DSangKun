from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = ROOT / 'frontend' / 'src' / 'constants' / 'dailyReportSheets.ts'


def normalize_cell(value: object) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def sheet_to_rows(workbook_path: Path):
    workbook = load_workbook(workbook_path, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows: list[list[str]] = []
            max_cols = 0

            for row in sheet.iter_rows(values_only=True):
                values = [normalize_cell(cell) for cell in row]
                while values and values[-1] == '':
                    values.pop()
                rows.append(values)
                max_cols = max(max_cols, len(values))

            normalized_rows: list[list[str]] = []
            for row in rows:
                padded = row + [''] * (max_cols - len(row))
                if any(cell != '' for cell in padded):
                    normalized_rows.append(padded)

            if not normalized_rows:
                continue

            yield {
                'sheetName': sheet.title,
                'sourceFile': workbook_path.name,
                'rows': normalized_rows
            }
    finally:
        workbook.close()


def collect_from_files(files: Iterable[Path]):
    collected: dict[str, dict] = {}

    for file_path in files:
        for sheet in sheet_to_rows(file_path):
            collected[sheet['sheetName']] = sheet

    return dict(sorted(collected.items(), key=lambda item: item[0]))


def to_ts_module(payload: dict[str, dict]) -> str:
    json_body = json.dumps(payload, ensure_ascii=False, indent=2)
    return (
        "export type DailyReportSheet = {\n"
        "  sheetName: string\n"
        "  sourceFile: string\n"
        "  rows: string[][]\n"
        "}\n\n"
        f"export const dailyReportSheets: Record<string, DailyReportSheet> = {json_body}\n"
    )


def main() -> int:
    parser = argparse.ArgumentParser(description='Extract all sheets from daily report Excel files into a TS constant module.')
    parser.add_argument('files', nargs='+', help='Excel files to extract (.xlsx)')
    parser.add_argument('--output', default=str(DEFAULT_OUTPUT), help='Output TypeScript file path')
    args = parser.parse_args()

    input_files = [Path(item).expanduser().resolve() for item in args.files]
    missing = [str(path) for path in input_files if not path.exists()]
    if missing:
        raise FileNotFoundError(f'Input files not found: {missing}')

    output_path = Path(args.output).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    payload = collect_from_files(input_files)
    output_path.write_text(to_ts_module(payload), encoding='utf-8')

    print(f'Generated {output_path}')
    print(f'Sheet count: {len(payload)}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
