#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
模块3离线对账脚本（不依赖 MySQL）

用途：
- 基于真实样本文件离线重建模块3关键指标
- 与最终统计报表做对账
- 当前优先验证单日 / 单子ASIN样本

当前重点口径：
1. 子ASIN卡片的 listing转化率：当前优先按“订单商品总数 / 会话数 - 总计”对齐
2. 子ASIN卡片里的广告订单 / 广告点击 / 花费等：当前样本更接近 SP 口径
3. 父组块顶部的总广告订单 / 总自然流量 / 自然转化率：按父组汇总口径重建

依赖：
  pip install pandas openpyxl

示例：
  python ops/reconcile_module3_summary.py ^
    --mapping-file D:\Programs\testdata\流量数据统计ASIN映射表.xlsx ^
    --biz-child-path D:\Programs\testdata\EXPERLAM-美国\20260317 ^
    --biz-parent-path D:\Programs\testdata\EXPERLAM-美国\20260317 ^
    --sp-path D:\Programs\testdata\EXPERLAM-美国\20260317 ^
    --sbv-path D:\Programs\testdata\SBV数据\20260317 ^
    --final-report D:\Programs\testdata\EXPERLAM-美国\20260317\EXPERLAM-美国-流量数据统计报表-20260317.xlsx ^
    --child-asin B0D45HSHDF
"""

import argparse
import datetime as dt
import json
import math
import re
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook

DATE_PATTERN = re.compile(r'(20\d{6})')

MAPPING_COLUMNS = [
    '父ASIN',
    '父ASIN产品名',
    '子ASIN',
    '子ASIN产品名',
    '是否有SBV',
    '归属人',
]

CHILD_BIZ_COLUMNS = {
    'parent_asin': '（父）ASIN',
    'child_asin': '（子）ASIN',
    'title': '标题',
    'sessions_total': '会话数 - 总计',
    'conversion_rate_total': '转化率 - 总计',
    'unit_session_pct': '商品会话百分比',
    'ordered_product_sales': '已订购商品销售额',
    'units_ordered': '已订购商品数量',
    'total_order_items': '订单商品总数',
}

PARENT_BIZ_COLUMNS = {
    'parent_asin': '（父）ASIN',
    'title': '标题',
    'sessions_total': '会话数 - 总计',
    'conversion_rate_total': '转化率 - 总计',
    'unit_session_pct': '商品会话百分比',
    'ordered_product_sales': '已订购商品销售额',
    'units_ordered': '已订购商品数量',
    'total_order_items': '订单商品总数',
}

SP_COLUMNS = {
    'ad_asin': '广告ASIN',
    'clicks': '点击量',
    'spend': '花费',
    'ad_sales_7d': '7天总销售额',
    'ad_orders_7d': '7天总订单数(#)',
    'impressions': '展示量',
}

FINAL_PARENT_KEYS = {
    'total_orders': '总订单',
    'total_ad_orders': '总广告订单',
    'total_natural_orders': '总自然订单',
    'parent_total_sessions': '父体总流量',
    'sp_clicks': 'SP流量',
    'sbv_clicks': 'SBV流量',
    'natural_sessions': '总自然流量',
    'natural_conversion_rate': '自然转化率',
    'listing_conversion_rate': 'listing转化率',
}

FINAL_CHILD_KEYS = {
    'orders': '订单数',
    'units': '销量数',
    'ad_orders': '广告订单',
    'sessions_total': '总流量',
    'ad_clicks': '广告点击',
    'ad_impressions': '广告展示量',
    'ctr': '点击率',
    'spend': '花费',
    'cpc': 'CPC',
    'cvr': 'CVR',
    'acos': 'ACOS',
    'listing_conversion_rate': 'listing转化率',
}


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('--mapping-file', required=True)
    parser.add_argument('--biz-child-path', required=True)
    parser.add_argument('--biz-parent-path', required=True)
    parser.add_argument('--sp-path', required=True)
    parser.add_argument('--sbv-path', required=True)
    parser.add_argument('--final-report', required=True)
    parser.add_argument('--child-asin', required=True)
    parser.add_argument('--stat-date', default=None, help='YYYY-MM-DD；缺省时自动从文件名提取')
    parser.add_argument('--json', action='store_true', help='输出 JSON 结果')
    return parser.parse_args()


def normalize_text(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    text = str(value).strip()
    if text == '' or text.lower() == 'nan':
        return None
    return text


def normalize_asin(value) -> Optional[str]:
    text = normalize_text(value)
    return text.upper() if text else None


def parse_bool_flag(value) -> int:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0
    text = str(value).strip().lower()
    return 1 if text in {'1', '1.0', 'true', 'yes', 'y', '是'} else 0


def parse_int(value, default: int = 0) -> int:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return default
    text = str(value).strip().replace(',', '')
    if text == '' or text == '--':
        return default
    try:
        return int(float(text))
    except Exception:
        return default


def parse_float(value, default: float = 0.0) -> float:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return default
    text = str(value).strip().replace(',', '')
    text = re.sub(r'^[A-Za-z$¥€£]+', '', text)
    if text == '' or text == '--':
        return default
    try:
        return float(text)
    except Exception:
        return default


def parse_percent_to_decimal(value) -> Optional[float]:
    text = normalize_text(value)
    if text is None or text == '--':
        return None
    if text.endswith('%'):
        try:
            return float(text[:-1].replace(',', '')) / 100.0
        except Exception:
            return None
    try:
        numeric = float(text.replace(',', ''))
        return numeric if numeric <= 1 else numeric / 100.0
    except Exception:
        return None


def parse_stat_date(name: str) -> dt.date:
    match = DATE_PATTERN.search(name)
    if not match:
        raise ValueError(f'无法从文件名提取日期: {name}')
    return dt.datetime.strptime(match.group(1), '%Y%m%d').date()


def safe_ratio(numerator: float, denominator: float) -> Optional[float]:
    if denominator in (None, 0):
        return None
    return numerator / denominator


def find_one(path_value: str, suffixes: Sequence[str], keywords: Iterable[str], stat_date: Optional[dt.date]) -> Path:
    path = Path(path_value)
    if not path.exists():
        raise FileNotFoundError(f'路径不存在: {path}')
    if path.is_file():
        return path

    candidates: List[Path] = []
    for suffix in suffixes:
        for p in path.rglob(f'*{suffix}'):
            if keywords and not any(k in p.name for k in keywords):
                continue
            if stat_date and DATE_PATTERN.search(p.name):
                if parse_stat_date(p.name) != stat_date:
                    continue
            candidates.append(p)

    if not candidates:
        raise FileNotFoundError(f'未找到匹配文件: path={path_value}, keywords={list(keywords)}')
    return sorted(candidates)[0]


def load_mapping(mapping_file: Path) -> pd.DataFrame:
    df = pd.read_excel(mapping_file)
    missing = [c for c in MAPPING_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f'映射表缺少字段: {missing}')
    df = df[MAPPING_COLUMNS].copy()
    df['父ASIN'] = df['父ASIN'].ffill()
    df['父ASIN产品名'] = df['父ASIN产品名'].ffill()
    df['子ASIN'] = df['子ASIN'].apply(normalize_asin)
    df['父ASIN'] = df['父ASIN'].apply(normalize_asin)
    return df[df['子ASIN'].notna()].copy()


def get_target_context(mapping_df: pd.DataFrame, child_asin: str) -> Dict:
    row = mapping_df[mapping_df['子ASIN'] == child_asin]
    if row.empty:
        raise ValueError(f'映射表中未找到子ASIN: {child_asin}')
    row = row.iloc[0]
    parent_asin = normalize_asin(row['父ASIN'])
    parent_name = normalize_text(row['父ASIN产品名'])
    child_name = normalize_text(row['子ASIN产品名']) or child_asin
    siblings = mapping_df[mapping_df['父ASIN'] == parent_asin].copy()
    siblings['has_sbv'] = siblings['是否有SBV'].apply(parse_bool_flag)
    return {
        'parent_asin': parent_asin,
        'parent_name': parent_name,
        'child_asin': child_asin,
        'child_name': child_name,
        'children': [a for a in siblings['子ASIN'].tolist() if a],
        'sbv_children': [
            normalize_asin(v)
            for v in siblings.loc[siblings['has_sbv'] == 1, '子ASIN'].tolist()
            if normalize_asin(v)
        ],
    }


def load_child_biz(file_path: Path, child_asin: str) -> Dict:
    df = pd.read_csv(file_path)
    row = df[df[CHILD_BIZ_COLUMNS['child_asin']].astype(str).str.upper() == child_asin]
    if row.empty:
        raise ValueError(f'子业务报告中未找到子ASIN: {child_asin}')
    row = row.iloc[0]
    return {
        'parent_asin': normalize_asin(row[CHILD_BIZ_COLUMNS['parent_asin']]),
        'title': normalize_text(row[CHILD_BIZ_COLUMNS['title']]),
        'sessions_total': parse_int(row[CHILD_BIZ_COLUMNS['sessions_total']]),
        'conversion_rate_total': parse_percent_to_decimal(row[CHILD_BIZ_COLUMNS['conversion_rate_total']]),
        'unit_session_pct': parse_percent_to_decimal(row[CHILD_BIZ_COLUMNS['unit_session_pct']]),
        'ordered_product_sales': parse_float(row[CHILD_BIZ_COLUMNS['ordered_product_sales']]),
        'units_ordered': parse_int(row[CHILD_BIZ_COLUMNS['units_ordered']]),
        'order_item_count': parse_int(row[CHILD_BIZ_COLUMNS['total_order_items']]),
    }


def load_parent_biz(file_path: Path, parent_asin: str) -> Dict:
    df = pd.read_csv(file_path)
    row = df[df[PARENT_BIZ_COLUMNS['parent_asin']].astype(str).str.upper() == parent_asin]
    if row.empty:
        raise ValueError(f'父业务报告中未找到父ASIN: {parent_asin}')
    row = row.iloc[0]
    return {
        'title': normalize_text(row[PARENT_BIZ_COLUMNS['title']]),
        'sessions_total': parse_int(row[PARENT_BIZ_COLUMNS['sessions_total']]),
        'conversion_rate_total': parse_percent_to_decimal(row[PARENT_BIZ_COLUMNS['conversion_rate_total']]),
        'unit_session_pct': parse_percent_to_decimal(row[PARENT_BIZ_COLUMNS['unit_session_pct']]),
        'ordered_product_sales': parse_float(row[PARENT_BIZ_COLUMNS['ordered_product_sales']]),
        'units_ordered': parse_int(row[PARENT_BIZ_COLUMNS['units_ordered']]),
        'order_item_count': parse_int(row[PARENT_BIZ_COLUMNS['total_order_items']]),
    }


def load_sp_metrics(file_path: Path, child_asin: str, group_children: Sequence[str]) -> Dict:
    df = pd.read_excel(file_path)
    target = df[df[SP_COLUMNS['ad_asin']].astype(str).str.upper() == child_asin]
    group = df[df[SP_COLUMNS['ad_asin']].astype(str).str.upper().isin(set(group_children))]
    return {
        'child': {
            'impressions': int(target[SP_COLUMNS['impressions']].fillna(0).sum()),
            'clicks': int(target[SP_COLUMNS['clicks']].fillna(0).sum()),
            'spend': float(target[SP_COLUMNS['spend']].fillna(0).sum()),
            'ad_sales': float(target[SP_COLUMNS['ad_sales_7d']].fillna(0).sum()),
            'ad_orders': int(target[SP_COLUMNS['ad_orders_7d']].fillna(0).sum()),
        },
        'group': {
            'impressions': int(group[SP_COLUMNS['impressions']].fillna(0).sum()),
            'clicks': int(group[SP_COLUMNS['clicks']].fillna(0).sum()),
            'spend': float(group[SP_COLUMNS['spend']].fillna(0).sum()),
            'ad_sales': float(group[SP_COLUMNS['ad_sales_7d']].fillna(0).sum()),
            'ad_orders': int(group[SP_COLUMNS['ad_orders_7d']].fillna(0).sum()),
        },
    }


def load_sbv_summary_row(file_path: Path) -> Dict:
    df = pd.read_excel(file_path)
    if df.empty:
        return {'clicks': 0, 'spend': 0.0, 'ad_sales': 0.0, 'ad_orders': 0, 'impressions': 0}

    first = df.iloc[0]
    is_summary_row = normalize_text(first.get('广告活动')) is None

    if is_summary_row:
        return {
            'impressions': parse_int(first.get('曝光量')),
            'clicks': parse_int(first.get('点击')),
            'spend': parse_float(first.get('花费')),
            'ad_sales': parse_float(first.get('广告销售额')),
            'ad_orders': parse_int(first.get('广告订单')),
        }

    return {
        'impressions': int(df.get('曝光量', pd.Series(dtype=float)).fillna(0).sum()),
        'clicks': int(df.get('点击', pd.Series(dtype=float)).fillna(0).sum()),
        'spend': float(df.get('花费', pd.Series(dtype=float)).fillna(0).sum()),
        'ad_sales': float(df.get('广告销售额', pd.Series(dtype=float)).fillna(0).sum()),
        'ad_orders': int(df.get('广告订单', pd.Series(dtype=float)).fillna(0).sum()),
    }


def load_sbv_metrics(sbv_dir: Path, child_asin: str, sbv_children: Sequence[str]) -> Dict:
    child_path = sbv_dir / f'{child_asin}.xlsx'
    child = load_sbv_summary_row(child_path) if child_path.exists() else {
        'impressions': 0, 'clicks': 0, 'spend': 0.0, 'ad_sales': 0.0, 'ad_orders': 0
    }

    group = {'impressions': 0, 'clicks': 0, 'spend': 0.0, 'ad_sales': 0.0, 'ad_orders': 0}
    for asin in sbv_children:
        path = sbv_dir / f'{asin}.xlsx'
        if not path.exists():
            continue
        item = load_sbv_summary_row(path)
        group['impressions'] += item['impressions']
        group['clicks'] += item['clicks']
        group['spend'] += item['spend']
        group['ad_sales'] += item['ad_sales']
        group['ad_orders'] += item['ad_orders']

    return {'child': child, 'group': group}


def parse_final_report_sections(final_report: Path, preferred_sheet: Optional[str] = None) -> Tuple[str, List[Dict[str, object]]]:
    wb = load_workbook(final_report, data_only=True)

    sheet_candidates = list(wb.sheetnames)
    if preferred_sheet and preferred_sheet in wb.sheetnames:
        sheet_candidates = [preferred_sheet] + [s for s in wb.sheetnames if s != preferred_sheet]

    for sheet_name in sheet_candidates:
        ws = wb[sheet_name]
        sections: List[Dict[str, object]] = []
        current_section_name = None
        current_metrics: Dict[str, object] = {}
        for row in ws.iter_rows(values_only=True):
            a, b, c = (row + (None, None, None))[:3]
            section_name = normalize_text(a)
            metric_name = normalize_text(b)
            value = c

            if section_name:
                if current_section_name is not None:
                    sections.append({'name': current_section_name, 'metrics': current_metrics})
                current_section_name = section_name
                current_metrics = {}
                if metric_name:
                    current_metrics[metric_name] = value
            elif current_section_name is not None and metric_name:
                current_metrics[metric_name] = value

        if current_section_name is not None:
            sections.append({'name': current_section_name, 'metrics': current_metrics})

        if sections:
            return sheet_name, sections

    raise ValueError(f'无法解析最终统计报表: {final_report}')


def pick_final_sections(final_report: Path, parent_name: str, child_name: str) -> Dict:
    wb = load_workbook(final_report, data_only=True)

    def _pick_from_sheet(sheet_name: str) -> Optional[Dict]:
        _, sections = parse_final_report_sections(final_report, preferred_sheet=sheet_name)
        if not sections:
            return None

        parent_section = sections[0]['metrics'] if sections[0]['name'] == parent_name else None
        child_section = None

        for idx, section in enumerate(sections):
            if idx == 0:
                continue
            if section['name'] == child_name:
                child_section = section['metrics']
                break

        if parent_section and child_section:
            return {'sheet_name': sheet_name, 'parent': parent_section, 'child': child_section}
        return None

    # 先优先用父产品名命中 sheet
    if parent_name in wb.sheetnames:
        matched = _pick_from_sheet(parent_name)
        if matched:
            return matched

    # 回退：扫描所有 sheet，找“首块=父组，后续块=子产品”的结构
    for sheet_name in wb.sheetnames:
        matched = _pick_from_sheet(sheet_name)
        if matched:
            return matched

    raise ValueError(f'最终统计报表中未找到 parent={parent_name}, child={child_name} 的对应区块')


def round_metric(value, digits: int = 4):
    if value is None:
        return None
    if isinstance(value, float):
        if math.isnan(value):
            return None
        return round(value, digits)
    return value


def format_percent(value: Optional[float]) -> Optional[str]:
    if value is None:
        return None
    return f'{value * 100:.2f}%'


def build_reconcile_result(context: Dict, child_biz: Dict, parent_biz: Dict, sp: Dict, sbv: Dict, final_sections: Dict, stat_date: dt.date) -> Dict:
    derived_child_listing_rate = (
        safe_ratio(child_biz['order_item_count'], child_biz['sessions_total'])
        if child_biz['sessions_total'] not in (None, 0)
        else child_biz['unit_session_pct']
        if child_biz['unit_session_pct'] is not None
        else child_biz['conversion_rate_total']
        if child_biz['conversion_rate_total'] is not None
        else safe_ratio(child_biz['units_ordered'], child_biz['sessions_total'])
    )

    group_ad_orders = sp['group']['ad_orders'] + sbv['group']['ad_orders']
    group_natural_orders = max(parent_biz['order_item_count'] - group_ad_orders, 0)
    group_natural_sessions = max(parent_biz['sessions_total'] - sp['group']['clicks'] - sbv['group']['clicks'], 0)
    group_natural_conversion_rate = safe_ratio(group_natural_orders, group_natural_sessions)
    group_listing_conversion_rate = safe_ratio(parent_biz['order_item_count'], parent_biz['sessions_total'])

    final_parent = final_sections['parent']
    final_child = final_sections['child']

    inferred_child_sp_sales_from_report = None
    spend = parse_float(final_child.get(FINAL_CHILD_KEYS['spend']), default=0.0)
    acos = parse_percent_to_decimal(final_child.get(FINAL_CHILD_KEYS['acos']))
    if spend and acos not in (None, 0):
        inferred_child_sp_sales_from_report = spend / acos

    result = {
        'stat_date': stat_date.isoformat(),
        'scope': {
            'parent_asin': context['parent_asin'],
            'parent_name': context['parent_name'],
            'child_asin': context['child_asin'],
            'child_name': context['child_name'],
            'final_report_sheet': final_sections['sheet_name'],
        },
        'derived': {
            'child_card': {
                'orders': child_biz['order_item_count'],
                'ad_orders_sp_only': sp['child']['ad_orders'],
                'sp_clicks': sp['child']['clicks'],
                'sp_spend': round_metric(sp['child']['spend'], 2),
                'sp_sales': round_metric(sp['child']['ad_sales'], 2),
                'listing_conversion_rate': round_metric(derived_child_listing_rate),
                'listing_conversion_rate_pct': format_percent(derived_child_listing_rate),
            },
            'parent_group': {
                'total_orders': parent_biz['order_item_count'],
                'total_ad_orders_sp_plus_sbv': group_ad_orders,
                'total_natural_orders': group_natural_orders,
                'parent_total_sessions': parent_biz['sessions_total'],
                'sp_clicks': sp['group']['clicks'],
                'sbv_clicks': sbv['group']['clicks'],
                'natural_sessions': group_natural_sessions,
                'natural_conversion_rate': round_metric(group_natural_conversion_rate),
                'natural_conversion_rate_pct': format_percent(group_natural_conversion_rate),
                'listing_conversion_rate': round_metric(group_listing_conversion_rate),
                'listing_conversion_rate_pct': format_percent(group_listing_conversion_rate),
            },
        },
        'final_report': {
            'child_card': {
                'orders': parse_int(final_child.get(FINAL_CHILD_KEYS['orders'])),
                'ad_orders': parse_int(final_child.get(FINAL_CHILD_KEYS['ad_orders'])),
                'ad_clicks': parse_int(final_child.get(FINAL_CHILD_KEYS['ad_clicks'])),
                'spend': round_metric(parse_float(final_child.get(FINAL_CHILD_KEYS['spend'])), 2),
                'listing_conversion_rate': round_metric(parse_percent_to_decimal(final_child.get(FINAL_CHILD_KEYS['listing_conversion_rate']))),
                'listing_conversion_rate_pct': normalize_text(final_child.get(FINAL_CHILD_KEYS['listing_conversion_rate'])),
                'sp_sales_inferred_from_spend_div_acos': round_metric(inferred_child_sp_sales_from_report, 2),
            },
            'parent_group': {
                'total_orders': parse_int(final_parent.get(FINAL_PARENT_KEYS['total_orders'])),
                'total_ad_orders': parse_int(final_parent.get(FINAL_PARENT_KEYS['total_ad_orders'])),
                'total_natural_orders': parse_int(final_parent.get(FINAL_PARENT_KEYS['total_natural_orders'])),
                'parent_total_sessions': parse_int(final_parent.get(FINAL_PARENT_KEYS['parent_total_sessions'])),
                'sp_clicks': parse_int(final_parent.get(FINAL_PARENT_KEYS['sp_clicks'])),
                'sbv_clicks': parse_int(final_parent.get(FINAL_PARENT_KEYS['sbv_clicks'])),
                'natural_sessions': parse_int(final_parent.get(FINAL_PARENT_KEYS['natural_sessions'])),
                'natural_conversion_rate': round_metric(parse_percent_to_decimal(final_parent.get(FINAL_PARENT_KEYS['natural_conversion_rate']))),
                'natural_conversion_rate_pct': normalize_text(final_parent.get(FINAL_PARENT_KEYS['natural_conversion_rate'])),
                'listing_conversion_rate': round_metric(parse_percent_to_decimal(final_parent.get(FINAL_PARENT_KEYS['listing_conversion_rate']))),
                'listing_conversion_rate_pct': normalize_text(final_parent.get(FINAL_PARENT_KEYS['listing_conversion_rate'])),
            },
        },
        'conclusions': [
            '子ASIN卡片 listing转化率 当前更接近 子业务报告 的 订单商品总数 / 会话数-总计；unit_session_pct 可作为无会话数时的回退参考。',
            '当前样本中，子ASIN卡片上的 广告订单 / 广告点击 / 花费 更接近 SP 口径，而不是 SP+SBV 叠加。',
            '父组块顶部的 listing转化率 应按 父业务报告中的 订单商品总数 / 会话数-总计 重建，即 总订单 / 父体总流量。',
            '当前样本中，父组块顶部的 总广告订单 / 总自然流量 / 自然转化率 可按 父业务报告 + 组内SP + 组内SBV 重建。',
        ],
    }

    result['checks'] = {
        'child_orders_match': result['derived']['child_card']['orders'] == result['final_report']['child_card']['orders'],
        'child_ad_orders_sp_match': result['derived']['child_card']['ad_orders_sp_only'] == result['final_report']['child_card']['ad_orders'],
        'child_sp_clicks_match': result['derived']['child_card']['sp_clicks'] == result['final_report']['child_card']['ad_clicks'],
        'child_sp_spend_match': abs(result['derived']['child_card']['sp_spend'] - result['final_report']['child_card']['spend']) < 0.01,
        'child_listing_conversion_rate_match': result['derived']['child_card']['listing_conversion_rate_pct'] == result['final_report']['child_card']['listing_conversion_rate_pct'],
        'parent_total_ad_orders_match': result['derived']['parent_group']['total_ad_orders_sp_plus_sbv'] == result['final_report']['parent_group']['total_ad_orders'],
        'parent_natural_sessions_match': result['derived']['parent_group']['natural_sessions'] == result['final_report']['parent_group']['natural_sessions'],
        'parent_natural_conversion_rate_match': result['derived']['parent_group']['natural_conversion_rate_pct'] == result['final_report']['parent_group']['natural_conversion_rate_pct'],
    }

    return result


def print_text_result(result: Dict):
    print('=== 模块3离线对账结果 ===')
    print(f"日期: {result['stat_date']}")
    print(f"父组: {result['scope']['parent_name']} ({result['scope']['parent_asin']})")
    print(f"子ASIN: {result['scope']['child_name']} ({result['scope']['child_asin']})")
    print(f"最终报表sheet: {result['scope']['final_report_sheet']}")
    print()

    print('--- 子ASIN卡片对账 ---')
    child_d = result['derived']['child_card']
    child_r = result['final_report']['child_card']
    print(f"订单数:        derived={child_d['orders']} | report={child_r['orders']}")
    print(f"广告订单:      derived(SP)={child_d['ad_orders_sp_only']} | report={child_r['ad_orders']}")
    print(f"SP点击:        derived={child_d['sp_clicks']} | report广告点击={child_r['ad_clicks']}")
    print(f"SP花费:        derived={child_d['sp_spend']:.2f} | report={child_r['spend']:.2f}")
    print(f"SP销售额:      derived={child_d['sp_sales']:.2f} | report推算={child_r['sp_sales_inferred_from_spend_div_acos']}")
    print(f"listing转化率: derived={child_d['listing_conversion_rate_pct']} | report={child_r['listing_conversion_rate_pct']}")
    print()

    print('--- 父组汇总对账 ---')
    parent_d = result['derived']['parent_group']
    parent_r = result['final_report']['parent_group']
    print(f"总订单:        derived={parent_d['total_orders']} | report={parent_r['total_orders']}")
    print(f"总广告订单:    derived(SP+SBV)={parent_d['total_ad_orders_sp_plus_sbv']} | report={parent_r['total_ad_orders']}")
    print(f"总自然订单:    derived={parent_d['total_natural_orders']} | report={parent_r['total_natural_orders']}")
    print(f"父体总流量:    derived={parent_d['parent_total_sessions']} | report={parent_r['parent_total_sessions']}")
    print(f"SP流量:        derived={parent_d['sp_clicks']} | report={parent_r['sp_clicks']}")
    print(f"SBV流量:       derived={parent_d['sbv_clicks']} | report={parent_r['sbv_clicks']}")
    print(f"自然流量:      derived={parent_d['natural_sessions']} | report={parent_r['natural_sessions']}")
    print(f"自然转化率:    derived={parent_d['natural_conversion_rate_pct']} | report={parent_r['natural_conversion_rate_pct']}")
    print(f"listing转化率: derived={parent_d['listing_conversion_rate_pct']} | report={parent_r['listing_conversion_rate_pct']}")
    print()

    print('--- 检查结论 ---')
    for key, value in result['checks'].items():
        print(f"{key}: {'PASS' if value else 'FAIL'}")
    print()

    print('--- 口径结论 ---')
    for line in result['conclusions']:
        print(f'- {line}')


def main():
    args = parse_args()
    child_asin = args.child_asin.strip().upper()

    stat_date = dt.date.fromisoformat(args.stat_date) if args.stat_date else None
    if stat_date is None:
        stat_date = parse_stat_date(Path(args.final_report).name)

    mapping_file = Path(args.mapping_file)
    child_biz_file = find_one(args.biz_child_path, ['.csv'], ['业务报告-子AINS', '业务报告-子ASIN'], stat_date)
    parent_biz_file = find_one(args.biz_parent_path, ['.csv'], ['业务报告-父AINS', '业务报告-父ASIN'], stat_date)
    sp_file = find_one(args.sp_path, ['.xlsx'], ['SP广告报告-'], stat_date)
    sbv_dir = Path(args.sbv_path)
    final_report = Path(args.final_report)

    mapping_df = load_mapping(mapping_file)
    context = get_target_context(mapping_df, child_asin)
    child_biz = load_child_biz(child_biz_file, child_asin)
    parent_biz = load_parent_biz(parent_biz_file, context['parent_asin'])
    sp = load_sp_metrics(sp_file, child_asin, context['children'])
    sbv = load_sbv_metrics(sbv_dir, child_asin, context['sbv_children'])
    final_sections = pick_final_sections(final_report, context['parent_name'], context['child_name'])

    result = build_reconcile_result(context, child_biz, parent_biz, sp, sbv, final_sections, stat_date)

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print_text_result(result)


if __name__ == '__main__':
    main()
