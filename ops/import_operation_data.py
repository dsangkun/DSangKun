#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
将 D:\Programs\testdata 下的运营 Excel 报表导入 MySQL。

依赖：
  pip install openpyxl pymysql

示例：
  python ops/import_operation_data.py \
    --host 127.0.0.1 --port 3306 --user root --password 123456 \
    --database ecommerce_ops --source-dir D:\\Programs\\testdata \
    --shop-name EXPERLAM --market-name 美国
"""

import argparse
import datetime as dt
import hashlib
import os
import re
import sys
from typing import Optional, Tuple

import pymysql
from openpyxl import load_workbook

DATE_PATTERN = re.compile(r'(20\d{6})')


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('--host', required=True)
    parser.add_argument('--port', type=int, default=3306)
    parser.add_argument('--user', required=True)
    parser.add_argument('--password', required=True)
    parser.add_argument('--database', required=True)
    parser.add_argument('--source-dir', required=True)
    parser.add_argument('--shop-name', required=True)
    parser.add_argument('--market-name', default='美国')
    parser.add_argument('--batch-code', default=None)
    parser.add_argument('--skip-existing', action='store_true', default=False)
    return parser.parse_args()


def md5_file(path: str) -> str:
    md5 = hashlib.md5()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(8192), b''):
            md5.update(chunk)
    return md5.hexdigest()


def parse_report_date(filename: str) -> dt.date:
    m = DATE_PATTERN.search(filename)
    if not m:
        raise ValueError(f'无法从文件名提取日期: {filename}')
    return dt.datetime.strptime(m.group(1), '%Y%m%d').date()


def normalize_metric_value(value) -> Tuple[Optional[str], Optional[float], Optional[str]]:
    if value is None:
        return None, None, None
    raw = str(value).strip()
    if raw == '':
        return '', None, None
    if raw.endswith('%'):
        try:
            return raw, float(raw[:-1]), 'PERCENT'
        except ValueError:
            return raw, None, 'TEXT'
    if isinstance(value, (int, float)):
        return raw, float(value), 'NUMBER'
    try:
        return raw, float(raw.replace(',', '')), 'NUMBER'
    except ValueError:
        return raw, None, 'TEXT'


def detect_entity_type(sheet_name: str, entity_name: str) -> str:
    if not entity_name:
        return 'UNKNOWN'
    if entity_name == sheet_name:
        return 'PARENT_SUMMARY'
    if entity_name.upper().endswith('SBV'):
        return 'CHILD_SBV'
    return 'CHILD_VARIANT'


def ensure_batch(conn, batch_code: str, source_dir: str) -> int:
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO op_import_batch(batch_code, source_dir, status)
            VALUES (%s, %s, 'RUNNING')
            ON DUPLICATE KEY UPDATE source_dir = VALUES(source_dir), status = 'RUNNING'
            """,
            (batch_code, source_dir),
        )
        conn.commit()
        cur.execute('SELECT id FROM op_import_batch WHERE batch_code = %s', (batch_code,))
        row = cur.fetchone()
        return row[0]


def upsert_report_file(conn, batch_id: int, shop_name: str, market_name: str, report_date: dt.date, path: str, skip_existing: bool) -> Tuple[int, bool]:
    filename = os.path.basename(path)
    file_md5 = md5_file(path)
    with conn.cursor() as cur:
        cur.execute('SELECT id FROM op_report_file WHERE source_filename = %s', (filename,))
        row = cur.fetchone()
        if row:
            if skip_existing:
                return row[0], True
            report_file_id = row[0]
            cur.execute('DELETE FROM op_metric_daily WHERE report_file_id = %s', (report_file_id,))
            cur.execute('DELETE FROM op_product_sheet WHERE report_file_id = %s', (report_file_id,))
            cur.execute(
                """
                UPDATE op_report_file
                   SET import_batch_id=%s, shop_name=%s, market_name=%s, report_date=%s,
                       source_path=%s, file_md5=%s, status='IMPORTED'
                 WHERE id=%s
                """,
                (batch_id, shop_name, market_name, report_date, path, file_md5, report_file_id),
            )
        else:
            cur.execute(
                """
                INSERT INTO op_report_file(
                  import_batch_id, shop_name, market_name, report_date,
                  source_filename, source_path, file_md5, status
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, 'IMPORTED')
                """,
                (batch_id, shop_name, market_name, report_date, filename, path, file_md5),
            )
            report_file_id = cur.lastrowid
        conn.commit()
        return report_file_id, False


def insert_product_sheet(conn, report_file_id: int, sheet_name: str) -> int:
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO op_product_sheet(report_file_id, sheet_name, product_name)
            VALUES (%s, %s, %s)
            """,
            (report_file_id, sheet_name, sheet_name),
        )
        conn.commit()
        return cur.lastrowid


def insert_metric(conn, report_file_id: int, product_sheet_id: int, stat_date: dt.date,
                  entity_type: str, entity_name: str, metric_name: str,
                  metric_value_raw: Optional[str], metric_value_num: Optional[float],
                  metric_unit: Optional[str], row_no: int):
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO op_metric_daily(
              report_file_id, product_sheet_id, stat_date, entity_type, entity_name,
              metric_name, metric_value_raw, metric_value_num, metric_unit, row_no
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                report_file_id, product_sheet_id, stat_date, entity_type, entity_name,
                metric_name, metric_value_raw, metric_value_num, metric_unit, row_no,
            ),
        )


def import_file(conn, path: str, batch_id: int, shop_name: str, market_name: str, skip_existing: bool):
    report_date = parse_report_date(os.path.basename(path))
    report_file_id, skipped = upsert_report_file(conn, batch_id, shop_name, market_name, report_date, path, skip_existing)
    if skipped:
        print(f'[SKIP] {os.path.basename(path)} 已存在')
        return

    wb = load_workbook(path, data_only=True)
    total_metrics = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        product_sheet_id = insert_product_sheet(conn, report_file_id, sheet_name)
        current_entity_name = None

        for row_no, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if len(row) < 3:
                continue
            col_a, col_b, col_c = row[0], row[1], row[2]
            if col_b is None:
                continue

            metric_name = str(col_b).strip()
            if metric_name == '' or metric_name == '日期':
                continue

            if col_a is not None and str(col_a).strip() != '':
                current_entity_name = str(col_a).strip()

            entity_name = current_entity_name or sheet_name
            entity_type = detect_entity_type(sheet_name, entity_name)
            raw, num, unit = normalize_metric_value(col_c)

            insert_metric(
                conn=conn,
                report_file_id=report_file_id,
                product_sheet_id=product_sheet_id,
                stat_date=report_date,
                entity_type=entity_type,
                entity_name=entity_name,
                metric_name=metric_name,
                metric_value_raw=raw,
                metric_value_num=num,
                metric_unit=unit,
                row_no=row_no,
            )
            total_metrics += 1

        conn.commit()

    print(f'[OK] {os.path.basename(path)} -> report_file_id={report_file_id}, metrics={total_metrics}')


def main():
    args = parse_args()
    batch_code = args.batch_code or f'operation-import-{dt.datetime.now().strftime("%Y%m%d%H%M%S")}'

    files = [
        os.path.join(args.source_dir, name)
        for name in os.listdir(args.source_dir)
        if name.lower().endswith('.xlsx')
    ]
    files.sort()

    if not files:
        print('未找到可导入的 xlsx 文件', file=sys.stderr)
        sys.exit(1)

    conn = pymysql.connect(
        host=args.host,
        port=args.port,
        user=args.user,
        password=args.password,
        database=args.database,
        charset='utf8mb4',
        autocommit=False,
    )

    try:
        batch_id = ensure_batch(conn, batch_code, args.source_dir)
        for path in files:
            import_file(conn, path, batch_id, args.shop_name, args.market_name, args.skip_existing)
        with conn.cursor() as cur:
            cur.execute("UPDATE op_import_batch SET status='SUCCESS' WHERE id=%s", (batch_id,))
        conn.commit()
        print(f'导入完成，batch_id={batch_id}')
    except Exception as exc:
        conn.rollback()
        try:
            with conn.cursor() as cur:
                cur.execute("UPDATE op_import_batch SET status='FAILED', remark=%s WHERE batch_code=%s", (str(exc)[:500], batch_code))
            conn.commit()
        except Exception:
            pass
        raise
    finally:
        conn.close()


if __name__ == '__main__':
    main()
